
import streamlit as st
import pandas as pd
from collections import defaultdict, Counter
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

COL_START = 1   # A
COL_END   = 11  # K
COLS      = [get_column_letter(c) for c in range(COL_START, COL_END+1)]
MAX_CANDIDATES_PER_ROW = 500

def normalize_value(v, trim_spaces=True, case_sensitive=True):
    if isinstance(v, str):
        s = v.strip() if trim_spaces else v
        return s if case_sensitive else s.lower()
    return v

def read_sheet_values(file, sheet_name=None, trim_spaces=True, case_sensitive=True):
    """read_only 모드로 값만 빠르게 읽는다 (메모리 절약)."""
    file.seek(0)
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = []
    for row_cells in ws.iter_rows(min_col=COL_START, max_col=COL_END):
        r = row_cells[0].row
        if r is None:
            continue
        orig = {}
        norm = {}
        empty_all = True
        for idx, cell in enumerate(row_cells):
            v = cell.value
            col = COLS[idx]
            orig[col] = v
            norm[col] = normalize_value(v, trim_spaces, case_sensitive)
            if v not in (None, ""):
                empty_all = False
        if not empty_all:
            rows.append({"_row": r, "orig": orig, "norm": norm})
    wb.close()
    return rows

def fill_signature(fill):
    if fill is None:
        return ("none",)
    pt = getattr(fill, "patternType", None)
    if not pt or str(pt).lower() == "none":
        return ("none",)
    def color_tuple(c):
        if c is None:
            return None
        return (
            getattr(c, "type", None),
            getattr(c, "rgb", None),
            getattr(c, "indexed", None),
            getattr(c, "theme", None),
            getattr(c, "tint", None),
        )
    return (
        str(pt).lower(),
        color_tuple(getattr(fill, "fgColor", None)),
        color_tuple(getattr(fill, "start_color", None)),
        color_tuple(getattr(fill, "end_color", None)),
    )

def read_sheet_fills(file, sheet_name=None, data_rows=None):
    """fill 정보 읽기. data_rows가 주어지면 해당 행 범위만 스캔한다."""
    file.seek(0)
    wb = load_workbook(file, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    if data_rows:
        row_set = set(r["_row"] for r in data_rows)
        max_data_row = max(row_set) if row_set else 1
    else:
        max_data_row = ws.max_row or 1
        row_set = None
    fills = {}
    for row_cells in ws.iter_rows(min_row=1, max_row=max_data_row,
                                   min_col=COL_START, max_col=COL_END):
        r = row_cells[0].row
        if row_set is not None and r not in row_set:
            continue
        for cell in row_cells:
            fills[(r, cell.column)] = fill_signature(cell.fill)
    return fills

def read_sheet_data(file, sheet_name=None, trim_spaces=True, case_sensitive=True, read_fills=False):
    """값 읽기 + 선택적 fill 읽기."""
    rows = read_sheet_values(file, sheet_name, trim_spaces, case_sensitive)
    if read_fills:
        fills = read_sheet_fills(file, sheet_name, data_rows=rows)
        return rows, fills
    return rows

def row_tuple(norm_row, cols):
    return tuple(norm_row[col] for col in cols)

def _count_eq(old_norm, new_norm, cols):
    return sum(1 for col in cols if old_norm.get(col) == new_norm.get(col))

def best_pairing(new_rows, old_rows, cols, progress_bar=None):
    """역인덱스 + row-by-row greedy 매칭."""
    if not new_rows or not old_rows:
        return [], list(range(len(old_rows))), list(range(len(new_rows)))

    col_val_to_old = defaultdict(set)
    for i, o in enumerate(old_rows):
        for col in cols:
            v = o["norm"].get(col)
            if v is not None and v != "":
                col_val_to_old[(col, v)].add(i)

    new_info = []
    total_new = len(new_rows)
    for j, n in enumerate(new_rows):
        if progress_bar and j % 200 == 0:
            progress_bar.progress(j / total_new * 0.4,
                                  text=f"후보 탐색 중... ({j}/{total_new})")
        candidates = set()
        for col in cols:
            v = n["norm"].get(col)
            if v is not None and v != "":
                candidates.update(col_val_to_old.get((col, v), set()))
        if len(candidates) > MAX_CANDIDATES_PER_ROW:
            scored = []
            for i in candidates:
                eq = _count_eq(old_rows[i]["norm"], n["norm"], cols)
                scored.append((eq, i))
            scored.sort(reverse=True)
            candidates = set(i for _, i in scored[:MAX_CANDIDATES_PER_ROW])
        best_eq = 0
        for i in candidates:
            eq = _count_eq(old_rows[i]["norm"], n["norm"], cols)
            best_eq = max(best_eq, eq)
        if candidates:
            new_info.append((best_eq, j, candidates))

    new_info.sort(reverse=True)
    used_old, used_new = set(), set()
    pairs = []
    total_info = len(new_info)
    for idx, (_, j, candidates) in enumerate(new_info):
        if progress_bar and idx % 200 == 0:
            progress_bar.progress(0.4 + idx / max(total_info, 1) * 0.5,
                                  text=f"매칭 중... ({idx}/{total_info})")
        n = new_rows[j]
        best_eq, best_i = 0, -1
        for i in candidates:
            if i in used_old:
                continue
            eq = _count_eq(old_rows[i]["norm"], n["norm"], cols)
            if eq > best_eq:
                best_eq = eq
                best_i = i
        if best_i >= 0:
            pairs.append((best_i, j, best_eq))
            used_old.add(best_i)
            used_new.add(j)

    remaining_old = [i for i in range(len(old_rows)) if i not in used_old]
    remaining_new = [j for j in range(len(new_rows)) if j not in used_new]
    if remaining_old and remaining_new and len(remaining_old) * len(remaining_new) <= 10000:
        if progress_bar:
            progress_bar.progress(0.9, text="잔여 행 폴백 매칭 중...")
        fb = []
        for i in remaining_old:
            for j in remaining_new:
                eq = _count_eq(old_rows[i]["norm"], new_rows[j]["norm"], cols)
                if eq > 0:
                    fb.append((eq, i, j))
        fb.sort(reverse=True)
        for eq, i, j in fb:
            if i in used_old or j in used_new:
                continue
            pairs.append((i, j, eq))
            used_old.add(i)
            used_new.add(j)

    if progress_bar:
        progress_bar.progress(1.0, text="매칭 완료")

    leftover_old = [i for i in range(len(old_rows)) if i not in used_old]
    leftover_new = [j for j in range(len(new_rows)) if j not in used_new]
    return pairs, leftover_old, leftover_new

def build_diff_record(old_row, new_row, cols):
    """비교 대상 열(cols)의 변경만 표시한다."""
    changes = []
    for col in cols:
        ov = old_row["orig"].get(col)
        nv = new_row["orig"].get(col)
        if old_row["norm"].get(col) != new_row["norm"].get(col):
            changes.append(f"{col}열 '{ov}'→'{nv}'")
    msg = "; ".join(changes) if changes else "값 동일 (정규화 기준)"
    return {
        "기준행": old_row["_row"],
        "비교행": new_row["_row"],
        "변경요약": msg
    }

# ── UI ─────────────────────────────────────────────

st.set_page_config(page_title="엑셀 행 재정렬 안전 비교 (A~K)", layout="wide")
st.title("📘 엑셀 행 재정렬 안전 비교 (A~K)")
st.caption("기준 엑셀을 먼저 업로드해 **행 내용을 저장**하고, "
           "비교 엑셀을 올려 **정렬/순서 변경과 무관하게** 변경사항을 검출합니다.")

with st.expander("비교 옵션", expanded=True):
    trim_spaces = st.checkbox("앞뒤 공백 무시", value=True)
    case_sensitive = st.checkbox("대소문자 구분", value=True)
    exclude_rows_if_fill_changed = st.checkbox("색상(채우기) 변경된 행 제외", value=False)
    exclude_cols = st.multiselect(
        "비교에서 제외할 열 선택",
        options=COLS,
        default=[],
        help="선택한 열은 동일 여부 판단에서 제외됩니다. 데이터는 그대로 읽지만 비교 시 무시합니다.")
    compare_cols = [c for c in COLS if c not in exclude_cols]
    if exclude_cols:
        st.caption(f"비교 대상 열: **{', '.join(compare_cols)}** ({len(compare_cols)}개)")

st.subheader("1) 기준(이전) 파일 저장")
c1, c2 = st.columns(2)
with c1:
    file_old = st.file_uploader("기준 엑셀 파일", type=["xlsx"], key="old")
with c2:
    sheet_old = None
    if file_old:
        try:
            file_old.seek(0)
            wb = load_workbook(file_old, read_only=True, data_only=True)
            sheet_old = st.selectbox("시트 선택(기준)", options=wb.sheetnames, index=0)
            wb.close()
        except Exception as e:
            st.error(f"기준 파일 시트 읽기 실패: {e}")

if st.button("기준 데이터 저장", type="primary", disabled=not (file_old and sheet_old)):
    try:
        with st.spinner("기준 데이터 읽는 중..."):
            if exclude_rows_if_fill_changed:
                old_rows, old_fills = read_sheet_data(
                    file_old, sheet_old, trim_spaces, case_sensitive, read_fills=True)
                st.session_state["old_fills"] = old_fills
            else:
                old_rows = read_sheet_values(file_old, sheet_old, trim_spaces, case_sensitive)
            st.session_state["old_rows"] = old_rows
        st.success(f"기준 데이터 저장 완료: {len(old_rows)} 행")
    except Exception as e:
        st.exception(e)

st.subheader("2) 비교(이후) 파일 분석")
c3, c4 = st.columns(2)
with c3:
    file_new = st.file_uploader("비교 엑셀 파일", type=["xlsx"], key="new")
with c4:
    sheet_new = None
    if file_new:
        try:
            file_new.seek(0)
            wb2 = load_workbook(file_new, read_only=True, data_only=True)
            sheet_new = st.selectbox("시트 선택(비교)", options=wb2.sheetnames, index=0)
            wb2.close()
        except Exception as e:
            st.error(f"비교 파일 시트 읽기 실패: {e}")

if st.button("변경 사항 분석 실행", type="primary",
             disabled=not (file_new and sheet_new and ("old_rows" in st.session_state))):
    try:
        old_rows = st.session_state["old_rows"]

        with st.spinner("비교 파일 읽는 중..."):
            if exclude_rows_if_fill_changed:
                new_rows, new_fills = read_sheet_data(
                    file_new, sheet_new, trim_spaces, case_sensitive, read_fills=True)
                old_fills = st.session_state.get("old_fills", {})
                if not old_fills:
                    old_fills = read_sheet_fills(
                        file_old, sheet_old, data_rows=old_rows) if (file_old and sheet_old) else {}
                max_row_new = max([r["_row"] for r in new_rows] or [0])
                fill_changed_rows = set()
                for r in range(1, max_row_new + 1):
                    for c in range(COL_START, COL_END + 1):
                        if old_fills.get((r, c)) != new_fills.get((r, c)):
                            fill_changed_rows.add(r)
                            break
            else:
                new_rows = read_sheet_values(file_new, sheet_new, trim_spaces, case_sensitive)
                fill_changed_rows = set()

        excluded_label = f" (제외 열: {', '.join(exclude_cols)})" if exclude_cols else ""
        st.info(f"기준 {len(old_rows)}행, 비교 {len(new_rows)}행 읽기 완료. "
                f"비교 대상 {len(compare_cols)}열{excluded_label}로 분석 시작...")

        # 현재 compare_cols로 멀티셋 구축 (열 제외 옵션 즉시 반영)
        old_multiset = Counter([row_tuple(r["norm"], compare_cols) for r in old_rows])
        old_mapping = defaultdict(list)
        for idx, r in enumerate(old_rows):
            old_mapping[row_tuple(r["norm"], compare_cols)].append(idx)

        remaining_old_indices = set(range(len(old_rows)))
        remaining_new_indices = set(range(len(new_rows)))
        exact_pairs = []
        temp_multiset = old_multiset.copy()
        temp_mapping = {k: v.copy() for k, v in old_mapping.items()}

        for j, nr in enumerate(new_rows):
            if nr["_row"] in fill_changed_rows:
                continue
            t = row_tuple(nr["norm"], compare_cols)
            if temp_multiset.get(t, 0) > 0:
                i = temp_mapping[t].pop(0)
                temp_multiset[t] -= 1
                exact_pairs.append((i, j))
                remaining_old_indices.discard(i)
                remaining_new_indices.discard(j)

        old_left = [old_rows[i] for i in sorted(remaining_old_indices)]
        new_left = [new_rows[j] for j in sorted(remaining_new_indices)
                    if new_rows[j]["_row"] not in fill_changed_rows]

        if old_left and new_left:
            prog_bar = st.progress(0, text="부분 매칭 진행 중...")
            pairs, leftover_old_idx, leftover_new_idx = best_pairing(
                new_left, old_left, compare_cols, progress_bar=prog_bar)
            prog_bar.empty()
        else:
            pairs, leftover_old_idx, leftover_new_idx = [], [], []

        best_pairs = []
        sorted_old_left = sorted(remaining_old_indices)
        sorted_new_left = [j for j in sorted(remaining_new_indices)
                           if new_rows[j]["_row"] not in fill_changed_rows]
        for eq, i, j in sorted([(p[2], p[0], p[1]) for p in pairs], reverse=True):
            old_idx_global = sorted_old_left[i]
            new_idx_global = sorted_new_left[j]
            best_pairs.append((old_idx_global, new_idx_global, eq))

        unchanged_records = [{
            "기준행": old_rows[i]["_row"],
            "비교행": new_rows[j]["_row"],
            "상태": "동일(재정렬만)"
        } for i, j in exact_pairs]

        changes_records = []
        for i, j, eq in best_pairs:
            rec = build_diff_record(old_rows[i], new_rows[j], compare_cols)
            rec["일치열수"] = eq
            rec["상태"] = "변경"
            changes_records.append(rec)

        used_old = set([i for i, _, _ in best_pairs] + [i for i, _ in exact_pairs])
        used_new = set([j for _, j, _ in best_pairs] + [j for _, j in exact_pairs])

        removed_records = [{"기준행": old_rows[i]["_row"], "상태": "제거됨"}
                           for i in range(len(old_rows)) if i not in used_old]
        added_records = [{"비교행": new_rows[j]["_row"], "상태": "추가됨"}
                         for j in range(len(new_rows))
                         if (j not in used_new and new_rows[j]["_row"] not in fill_changed_rows)]

        df_unchanged = pd.DataFrame(unchanged_records)
        df_changes = pd.DataFrame(changes_records,
                                  columns=["기준행", "비교행", "일치열수", "변경요약", "상태"])
        df_removed = pd.DataFrame(removed_records)
        df_added = pd.DataFrame(added_records)

        st.success(
            f"결과: 동일(재정렬만) {len(df_unchanged)}건, "
            f"변경 {len(df_changes)}건, "
            f"제거 {len(df_removed)}건, "
            f"추가 {len(df_added)}건")
        st.write("### 동일(재정렬만)")
        st.dataframe(df_unchanged, use_container_width=True, hide_index=True)
        st.write("### 변경")
        st.dataframe(df_changes, use_container_width=True, hide_index=True)
        st.write("### 제거됨(기준에는 있었으나 비교에는 없음)")
        st.dataframe(df_removed, use_container_width=True, hide_index=True)
        st.write("### 추가됨(비교에는 있으나 기준에는 없음)")
        st.dataframe(df_added, use_container_width=True, hide_index=True)

        def to_xlsx(dfs, names):
            bio = BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as wr:
                for df, name in zip(dfs, names):
                    if not df.empty:
                        df.to_excel(wr, index=False, sheet_name=name)
                    else:
                        pd.DataFrame().to_excel(wr, index=False, sheet_name=name)
            return bio.getvalue()

        st.download_button(
            "결과 통합 엑셀 다운로드",
            data=to_xlsx(
                [df_unchanged, df_changes, df_removed, df_added],
                ["unchanged", "changes", "removed", "added"]),
            file_name="diff_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.exception(e)

st.info("주의: '기준 데이터 저장'을 먼저 수행하세요. "
        "그 다음 비교 파일을 업로드하고 '변경 사항 분석 실행'을 누르면, "
        "행 순서가 달라도 정확히 식별합니다.")
